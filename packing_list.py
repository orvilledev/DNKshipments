"""Parse "Package Content List" export workbooks into a flat box-contents table."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.styles import Font

CARTON_LABEL = "carton no:"
CARTON_TOTAL_LABEL = "carton total:"
ORDER_LABEL = "order no."
GRAND_TOTAL_LABEL = "total # of cartons:"
GRAND_QTY_LABEL = "total quantity:"
NET_WEIGHT_LABEL = "total net weight:"

ITEM = "item"
UPC = "upc"
SIZE = "size"
DESCRIPTION = "description"
QUANTITY = "quantity"
UOM = "uom"
UNIT_WEIGHT = "unit_weight"
TOTAL_WEIGHT = "total_weight"

HEADER_ALIASES = {
    "item": ITEM,
    "item no": ITEM,
    "item no.": ITEM,
    "item number": ITEM,
    "upc": UPC,
    "upcs": UPC,
    "upc code": UPC,
    "upc#": UPC,
    "gtin": UPC,
    "ean": UPC,
    "barcode": UPC,
    "size": SIZE,
    "description": DESCRIPTION,
    "quantity": QUANTITY,
    "qty": QUANTITY,
    "unit of measure code": UOM,
    "unit of measure": UOM,
    "uom": UOM,
    "weight": UNIT_WEIGHT,
    "unit weight": UNIT_WEIGHT,
    "wt": UNIT_WEIGHT,
    "total weight": TOTAL_WEIGHT,
    "totalweight": TOTAL_WEIGHT,
    "ext weight": TOTAL_WEIGHT,
    "extended weight": TOTAL_WEIGHT,
}

# Optional columns that often appear only on the first carton header, then keep
# the same column for the rest of the sheet even when the label is omitted.
# UPC may sit in column O, P, or elsewhere depending on the export — wherever
# the "UPC" header is found is treated as the source of truth, and if a later
# carton reprints "UPC" in a different column the sticky mapping moves with it.
# WEIGHT / TOTAL WEIGHT behave the same way.
STICKY_FIELDS = {UPC, UNIT_WEIGHT, TOTAL_WEIGHT}

CARTON_SEQ_RE = re.compile(r"carton:\s*(\d+)\s*of\s*(\d+)", re.IGNORECASE)
DIMENSIONS_RE = re.compile(r"dimensions:\s*(.+)", re.IGNORECASE)
NUMBER_RE = re.compile(r"\d+(?:[.,]\d+)?")

OUTPUT_COLUMNS = ["UPCs", "Box Number", "Quantity"]
DIMENSION_COLUMNS = ["Box Number", "Length", "Width", "Height", "Weight"]


class PackingListError(Exception):
    """Raised when a workbook does not look like a Package Content List export."""


@dataclass
class Carton:
    box_number: int
    carton_no: str = ""
    dimensions: str = ""
    order_no: str = ""
    reported_total: int | None = None
    sheet: str = ""
    length: float | None = None
    width: float | None = None
    height: float | None = None
    lines: list[dict[str, Any]] = field(default_factory=list)

    @property
    def parsed_total(self) -> int:
        return sum(line["Quantity"] for line in self.lines)

    @property
    def total_weight(self) -> float | None:
        """Sum of each line's TOTAL WEIGHT in this carton, if any were present."""
        weights = [
            line["Total Weight"]
            for line in self.lines
            if line.get("Total Weight") is not None
        ]
        if not weights:
            return None
        return float(sum(weights))


@dataclass
class ParsedPackingList:
    cartons: list[Carton]
    lines: pd.DataFrame
    reported_carton_count: int | None = None
    reported_total_quantity: int | None = None
    reported_net_weight: float | None = None
    order_numbers: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def total_quantity(self) -> int:
        return int(self.lines["Quantity"].sum()) if not self.lines.empty else 0

    @property
    def total_weight(self) -> float | None:
        if self.lines.empty or "Total Weight" not in self.lines.columns:
            return None
        weights = self.lines["Total Weight"].dropna()
        if weights.empty:
            return None
        return float(weights.sum())


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _display_number(value: Any, number_format: str | None) -> str:
    """Render a cell the way Excel shows it, honouring zero-padded formats.

    Item numbers such as 50020202 are stored as integers but carry a "000000000"
    format, so the printed packing list shows 050020202. Dropping that leading
    zero would corrupt the UPC.
    """
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    if not isinstance(value, int):
        return _text(value)

    fmt = (number_format or "").split(";")[0].strip().strip('"')
    if fmt and set(fmt) <= {"0"}:
        return str(abs(value)).zfill(len(fmt)) if value >= 0 else str(value)
    return str(value)


def parse_dimensions(text: str) -> tuple[float | None, float | None, float | None]:
    """Split a dimensions string such as "31x19x14" into length, width, height.

    The three numbers are read in printed order, so 31x19x14 is length 31,
    width 19, height 14. Separators other than "x" and trailing units are
    tolerated.
    """
    if not text:
        return (None, None, None)

    normalised = re.sub(r"[×✕✖*]", "x", str(text))
    parts = re.split(r"x", normalised, flags=re.IGNORECASE)
    numbers = [NUMBER_RE.search(part) for part in parts]
    values = [float(m.group(0).replace(",", ".")) for m in numbers if m]

    if len(values) < 3:
        # Fall back to any three numbers in the string, e.g. "L31 W19 H14".
        values = [
            float(m.group(0).replace(",", "."))
            for m in NUMBER_RE.finditer(normalised)
        ]

    if len(values) < 3:
        return (None, None, None)
    return (values[0], values[1], values[2])


def upc_display_widths(upcs: Any) -> dict[int, int]:
    """Map a UPC's numeric value to the digit count it is printed with.

    Only values with a leading zero need this: 050020202 is stored as the
    number 50020202, so Excel needs a "000000000" format to print it the way
    the packing list does.
    """
    widths: dict[int, int] = {}
    for upc in upcs:
        text = str(upc).strip()
        if text.isdigit() and text.startswith("0"):
            value = int(text)
            widths[value] = max(widths.get(value, 0), len(text))
    return widths


def _numeric_if_possible(series: pd.Series) -> pd.Series:
    """Convert a column of digit strings to real numbers, or leave it alone.

    Anything that is not numeric all the way down (sizes like "39-40", carton
    numbers like L00011021774) is left as text.
    """
    text = series.astype(str).str.strip()
    filled = text.ne("")
    if not filled.any():
        return series
    if text.str.match(r"0\d").any():
        # A leading zero is meaningful, so keep the printed form.
        return series

    numeric = pd.to_numeric(text.where(filled), errors="coerce")
    if numeric[filled].isna().any():
        return series
    if (numeric.dropna() % 1 == 0).all():
        return numeric.astype("Int64")
    return numeric


def _to_int(value: Any) -> int | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(round(value))
    text = _text(value).replace(",", "")
    if not text:
        return None
    try:
        return int(round(float(text)))
    except ValueError:
        return None


def _to_float(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _text(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _row_cells(row: tuple[Any, ...]) -> list[tuple[int, Any]]:
    """Non-empty cells of a row as (column index, cell) pairs."""
    return [(cell.column, cell) for cell in row if _text(cell.value) != ""]


def _row_labels(cells: list[tuple[int, Any]]) -> dict[int, str]:
    return {col: _text(cell.value).lower() for col, cell in cells}


def _find_label(labels: dict[int, str], needle: str) -> int | None:
    for col, text in labels.items():
        if text == needle or text.startswith(needle):
            return col
    return None


def _header_map(cells: list[tuple[int, Any]]) -> tuple[dict[int, str], bool] | None:
    """Map the column of each recognised header to its canonical field name.

    Returns ``(headers, is_complete)``, where a complete row carries both Item
    and Quantity. Extra labels such as Weight are ignored so they do not stop
    the row from counting as a header. "Unit of Measure Code" is printed on its
    own row above the other headers, so partial rows are reported too. ``None``
    means the row is not a header row at all.
    """
    found: dict[int, str] = {}
    for col, cell in cells:
        key = _text(cell.value).lower().rstrip(":")
        field_name = HEADER_ALIASES.get(key)
        if field_name is not None:
            found[col] = field_name
    if not found:
        return None
    complete = ITEM in found.values() and QUANTITY in found.values()
    return found, complete


def _merge_headers(
    found: dict[int, str],
    pending: dict[int, str],
    sticky: dict[int, str],
) -> dict[int, str]:
    """Combine a new header row with pending and sticky optional columns.

    Sticky columns (notably UPC) often appear only on the first carton header;
    later cartons drop the label but keep the values in the same column.
    """
    found_fields = set(found.values())
    pending_fields = set(pending.values())
    kept_sticky = {
        col: field
        for col, field in sticky.items()
        if col not in found
        and col not in pending
        and field not in found_fields
        and field not in pending_fields
    }
    return {**kept_sticky, **pending, **found}


def _update_sticky(sticky: dict[int, str], found: dict[int, str]) -> dict[int, str]:
    """Remember optional columns from a header so later cartons can reuse them."""
    updated = dict(sticky)
    for col, field in found.items():
        if field not in STICKY_FIELDS:
            continue
        updated = {c: f for c, f in updated.items() if f != field}
        updated[col] = field
    return updated


def _field_for_column(column: int, header_map: dict[int, str]) -> str | None:
    """Assign a value cell to a header.

    Values are not always left-aligned with their header (sizes sit one column
    right of the "Size" label), so each value belongs to the nearest header at
    or to the left of it.
    """
    candidates = [col for col in header_map if col <= column]
    if not candidates:
        return None
    return header_map[max(candidates)]


def _load_workbook(source: bytes | str | io.BytesIO) -> openpyxl.Workbook:
    if isinstance(source, bytes):
        source = io.BytesIO(source)
    try:
        return openpyxl.load_workbook(source, data_only=True)
    except Exception as exc:  # noqa: BLE001 - surfaced to the UI as a message
        raise PackingListError(
            "Could not read that workbook. Please upload the .xlsx Package "
            f"Content List export (openpyxl said: {exc})."
        ) from exc


def parse_packing_list(source: bytes | str | io.BytesIO) -> ParsedPackingList:
    workbook = _load_workbook(source)

    cartons: list[Carton] = []
    warnings: list[str] = []
    order_numbers: list[str] = []
    reported_carton_count: int | None = None
    reported_total_quantity: int | None = None
    reported_net_weight: float | None = None

    current: Carton | None = None
    header_map: dict[int, str] | None = None
    pending_headers: dict[int, str] = {}
    sticky_headers: dict[int, str] = {}
    current_order = ""

    for worksheet in workbook.worksheets:
        # Sticky optional columns are per-sheet; a fresh sheet may place UPC elsewhere.
        sticky_headers = {}
        for row in worksheet.iter_rows():
            cells = _row_cells(row)
            if not cells:
                continue
            labels = _row_labels(cells)
            row_number = cells[0][1].row

            order_col = _find_label(labels, ORDER_LABEL)
            if order_col is not None:
                value = next(
                    (_text(c.value) for col, c in cells if col > order_col), ""
                )
                if value:
                    current_order = value
                    if value not in order_numbers:
                        order_numbers.append(value)
                continue

            net_weight_col = _find_label(labels, NET_WEIGHT_LABEL)
            if net_weight_col is not None:
                reported_net_weight = next(
                    (
                        _to_float(c.value)
                        for col, c in cells
                        if col > net_weight_col and _to_float(c.value) is not None
                    ),
                    reported_net_weight,
                )
                # The same footer row also carries carton/quantity totals — keep going
                # so those labels on this row are still read below.

            total_col = _find_label(labels, GRAND_TOTAL_LABEL)
            if total_col is not None:
                reported_carton_count = next(
                    (
                        _to_int(c.value)
                        for col, c in cells
                        if col > total_col and _to_int(c.value) is not None
                    ),
                    reported_carton_count,
                )
                qty_col = _find_label(labels, GRAND_QTY_LABEL)
                if qty_col is not None:
                    reported_total_quantity = next(
                        (
                            _to_int(c.value)
                            for col, c in cells
                            if col > qty_col and _to_int(c.value) is not None
                        ),
                        reported_total_quantity,
                    )
                current = None
                continue

            carton_col = _find_label(labels, CARTON_LABEL)
            if carton_col is not None:
                carton_no = ""
                dimensions = ""
                seq: int | None = None
                for col, cell in cells:
                    if col <= carton_col:
                        continue
                    text = _text(cell.value)
                    match = CARTON_SEQ_RE.search(text)
                    if match:
                        seq = int(match.group(1))
                        continue
                    dims = DIMENSIONS_RE.search(text)
                    if dims:
                        dimensions = dims.group(1).strip()
                        continue
                    if not carton_no:
                        carton_no = text

                # A carton continued onto a new print page repeats its header;
                # keep appending to the same carton instead of duplicating it.
                if (
                    current is not None
                    and carton_no
                    and carton_no == current.carton_no
                    and (seq is None or seq == current.box_number)
                ):
                    continue

                box_number = seq if seq is not None else len(cartons) + 1
                if seq is not None and seq != len(cartons) + 1:
                    warnings.append(
                        f"Carton {carton_no or '(no number)'} is labelled "
                        f"'Carton: {seq}' but is block #{len(cartons) + 1} in the file."
                    )
                length, width, height = parse_dimensions(dimensions)
                current = Carton(
                    box_number=box_number,
                    carton_no=carton_no,
                    dimensions=dimensions,
                    order_no=current_order,
                    sheet=worksheet.title,
                    length=length,
                    width=width,
                    height=height,
                )
                cartons.append(current)
                header_map = None
                pending_headers = {}
                continue

            carton_total_col = _find_label(labels, CARTON_TOTAL_LABEL)
            if carton_total_col is not None:
                if current is not None:
                    current.reported_total = next(
                        (
                            _to_int(c.value)
                            for col, c in cells
                            if col > carton_total_col and _to_int(c.value) is not None
                        ),
                        None,
                    )
                header_map = None
                pending_headers = {}
                continue

            maybe_header = _header_map(cells)
            if maybe_header is not None:
                found, complete = maybe_header
                if complete:
                    header_map = _merge_headers(found, pending_headers, sticky_headers)
                    sticky_headers = _update_sticky(sticky_headers, found)
                    pending_headers = {}
                elif header_map is not None:
                    header_map.update(found)
                    sticky_headers = _update_sticky(sticky_headers, found)
                else:
                    pending_headers.update(found)
                    sticky_headers = _update_sticky(sticky_headers, found)
                continue

            if current is None or header_map is None:
                continue

            values: dict[str, Any] = {}
            for col, cell in cells:
                field_name = _field_for_column(col, header_map)
                if field_name is None or field_name in values:
                    continue
                if field_name in {ITEM, UPC}:
                    values[field_name] = _display_number(cell.value, cell.number_format)
                elif field_name == QUANTITY:
                    values[QUANTITY] = _to_int(cell.value)
                elif field_name in {UNIT_WEIGHT, TOTAL_WEIGHT}:
                    values[field_name] = _to_float(cell.value)
                else:
                    values[field_name] = _text(cell.value)

            # Prefer the dedicated UPC column when the packing list has one;
            # fall back to Item for older exports that only print Item.
            upc = _text(values.get(UPC)) or _text(values.get(ITEM))
            quantity = values.get(QUANTITY)
            if not upc:
                continue
            if quantity is None:
                warnings.append(
                    f"{worksheet.title}!row {row_number}: item {upc} has no quantity "
                    "and was skipped."
                )
                continue

            unit_weight = values.get(UNIT_WEIGHT)
            total_weight = values.get(TOTAL_WEIGHT)
            if total_weight is None and unit_weight is not None:
                total_weight = float(unit_weight) * int(quantity)

            current.lines.append(
                {
                    "UPCs": upc,
                    "Box Number": current.box_number,
                    "Quantity": int(quantity),
                    "Item": _text(values.get(ITEM)),
                    "Size": _text(values.get(SIZE)),
                    "Description": _text(values.get(DESCRIPTION)),
                    "Unit of Measure Code": _text(values.get(UOM)),
                    "Unit Weight": unit_weight,
                    "Total Weight": total_weight,
                    "Carton No": current.carton_no,
                    "Dimensions": current.dimensions,
                    "Order No": current.order_no,
                    "Source Row": f"{worksheet.title}!{row_number}",
                }
            )

    if not cartons:
        raise PackingListError(
            "No cartons were found. This app expects the 'Package Content List' "
            "export, which has rows starting with 'Carton No:'."
        )

    all_lines = [line for carton in cartons for line in carton.lines]
    if not all_lines:
        raise PackingListError(
            "Cartons were found but none of them contained item rows."
        )

    for carton in cartons:
        if carton.reported_total is not None and carton.reported_total != carton.parsed_total:
            warnings.append(
                f"Box {carton.box_number} ({carton.carton_no}): file says Carton "
                f"Total {carton.reported_total} but the item rows add up to "
                f"{carton.parsed_total}."
            )
        if carton.length is None:
            warnings.append(
                f"Box {carton.box_number} ({carton.carton_no}): could not read "
                + (
                    f"dimensions from '{carton.dimensions}'."
                    if carton.dimensions
                    else "dimensions, the carton row has none."
                )
            )

    lines = pd.DataFrame(all_lines)
    lines["Size"] = _numeric_if_possible(lines["Size"])
    return ParsedPackingList(
        cartons=cartons,
        lines=lines,
        reported_carton_count=reported_carton_count,
        reported_total_quantity=reported_total_quantity,
        reported_net_weight=reported_net_weight,
        order_numbers=order_numbers,
        warnings=warnings,
    )


def build_output(
    parsed: ParsedPackingList,
    combine_duplicates: bool = True,
    numeric_upcs: bool = True,
) -> pd.DataFrame:
    """Return the UPCs / Box Number / Quantity table.

    UPCs are grouped as printed strings, then turned into real numbers when
    every one of them is digits only. Zero-padded UPCs keep their printed form
    in Excel through a number format, see ``to_excel_bytes``.
    """
    output = parsed.lines[OUTPUT_COLUMNS].copy()
    if combine_duplicates:
        output = (
            output.groupby(["Box Number", "UPCs"], as_index=False, sort=False)["Quantity"]
            .sum()
            .loc[:, OUTPUT_COLUMNS]
        )
    output = output.sort_values(
        ["Box Number"], kind="stable", ignore_index=True
    )
    output["Box Number"] = output["Box Number"].astype(int)
    output["Quantity"] = output["Quantity"].astype(int)
    if numeric_upcs and output["UPCs"].astype(str).str.fullmatch(r"\d+").all():
        output["UPCs"] = output["UPCs"].astype("int64")
    return output


def build_dimensions(
    parsed: ParsedPackingList, include_box_number: bool = True
) -> pd.DataFrame:
    """Return one row per box with Length, Width, Height and total Weight."""
    rows = [
        {
            "Box Number": carton.box_number,
            "Length": carton.length,
            "Width": carton.width,
            "Height": carton.height,
            "Weight": carton.total_weight,
        }
        for carton in parsed.cartons
    ]
    dimensions = pd.DataFrame(rows, columns=DIMENSION_COLUMNS)
    dimensions = dimensions.drop_duplicates(subset=["Box Number"], keep="first")
    dimensions = dimensions.sort_values("Box Number", kind="stable", ignore_index=True)
    dimensions["Box Number"] = dimensions["Box Number"].astype(int)

    for column in ("Length", "Width", "Height", "Weight"):
        values = pd.to_numeric(dimensions[column], errors="coerce")
        present = values.dropna()
        # Whole numbers should read as 31, not 31.0.
        if not present.empty and (present % 1 == 0).all():
            values = values.astype("Int64")
        dimensions[column] = values

    if not include_box_number:
        dimensions = dimensions.drop(columns=["Box Number"])
    return dimensions


def to_excel_bytes(
    output: pd.DataFrame,
    dimensions: pd.DataFrame | None = None,
    detail: pd.DataFrame | None = None,
    sheet_name: str = "Box Contents",
    dimensions_sheet_name: str = "Box Dimensions",
    upc_widths: dict[int, int] | None = None,
) -> bytes:
    """Write the tables to a workbook.

    Numeric columns are written as numbers so Excel can sum and sort them
    without the "number stored as text" warning. ``upc_widths`` supplies the
    zero-padded display format for UPCs that are printed with a leading zero.
    """
    upc_widths = upc_widths or {}
    buffer = io.BytesIO()
    sheets: list[tuple[str, pd.DataFrame]] = [(sheet_name, output)]
    if dimensions is not None and not dimensions.empty:
        sheets.append((dimensions_sheet_name, dimensions))
    if detail is not None and not detail.empty:
        sheets.append(("Details", detail))

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for name, frame in sheets:
            frame.to_excel(writer, index=False, sheet_name=name)

        for name, frame in sheets:
            worksheet = writer.sheets[name]
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
            worksheet.freeze_panes = "A2"
            for index, column in enumerate(frame.columns, start=1):
                letter = openpyxl.utils.get_column_letter(index)
                width = max(len(str(column)), 12)
                if not frame.empty:
                    width = max(width, int(frame[column].astype(str).str.len().max()) + 2)
                worksheet.column_dimensions[letter].width = min(width, 46)

                series = frame[column]
                numeric = pd.api.types.is_numeric_dtype(series)
                cells = [
                    cell
                    for row in worksheet.iter_rows(min_row=2, min_col=index, max_col=index)
                    for cell in row
                ]
                if column == "UPCs" and numeric:
                    for cell in cells:
                        pad = upc_widths.get(cell.value) if isinstance(cell.value, int) else None
                        cell.number_format = "0" * pad if pad else "0"
                elif pd.api.types.is_integer_dtype(series):
                    for cell in cells:
                        cell.number_format = "0"
                elif not numeric:
                    # Text format stops Excel reading codes like "39-40" as dates.
                    for cell in cells:
                        cell.number_format = "@"
    return buffer.getvalue()
