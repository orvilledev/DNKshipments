"""Render app.py through Streamlit's AppTest, with and without an uploaded file."""

import io
import textwrap
from pathlib import Path

from streamlit.testing.v1 import AppTest

XLSX = Path(r"C:\Users\Administrator\Downloads\PMSH01 PO# FBA19HZZ856W.xlsx")

print("== no file uploaded ==")
at = AppTest.from_file("app.py", default_timeout=90).run()
assert not at.exception, at.exception
print("info:", [i.value for i in at.info])
print("toggles:", [(t.label, t.value) for t in at.toggle])

# AppTest cannot drive st.file_uploader, so stub it out and re-run the script.
harness = Path("_app_with_upload.py")
harness.write_text(
    textwrap.dedent(
        f"""
        import io
        import runpy
        import streamlit as st

        class _Uploaded(io.BytesIO):
            name = {XLSX.name!r}

        _data = open({str(XLSX)!r}, "rb").read()
        st.file_uploader = lambda *a, **k: _Uploaded(_data)
        runpy.run_path("app.py", run_name="__main__")
        """
    ).lstrip(),
    encoding="utf-8",
)

try:
    print("\n== with the real packing list uploaded ==")
    at = AppTest.from_file(str(harness), default_timeout=180).run()
    assert not at.exception, at.exception

    print("metrics:", [(m.label, m.value) for m in at.metric])
    print("success:", [s.value for s in at.success])
    print("warning:", [w.value for w in at.warning])
    print("error:", [e.value for e in at.error])
    print(
        "downloads:",
        [b.label.encode("ascii", "ignore").decode().strip() for b in at.get("download_button")],
    )
    print("dataframes:", [df.value.shape for df in at.dataframe])

    assert not at.error, "unexpected error message in the app"
    assert len(at.success) == 4, "expected all four reconciliation checks to pass"
    assert len(at.get("download_button")) == 3, "expected Excel + two CSV buttons"

    # Default is one row per printed line.
    main = at.dataframe[0].value
    assert list(main.columns) == ["UPCs", "Box Number", "Quantity"]
    assert main["Box Number"].nunique() == 28
    assert len(main) == 199 and int(main["Quantity"].sum()) == 333
    assert main["UPCs"].map(type).eq(str).all()
    box1 = main[main["Box Number"] == 1]
    print("\nbox 1 as printed:")
    print(box1.to_string(index=False))
    assert list(box1["Quantity"]) == [1, 3, 4, 2, 1, 1], "box 1 must match the printed carton"

    dims = at.dataframe[1].value
    print("\ndimensions tab:")
    print(dims.head(3).to_string(index=False))
    assert list(dims.columns) == ["Box Number", "Length", "Width", "Height"]
    assert len(dims) == 28
    assert list(dims.loc[0, ["Length", "Width", "Height"]]) == [31, 19, 14]

    print("\n== combine duplicates ON ==")
    at.sidebar.toggle[0].set_value(True).run()
    assert not at.exception, at.exception
    combined = at.dataframe[0].value
    print("rows:", combined.shape, "qty:", int(combined["Quantity"].sum()))
    assert len(combined) == 109 and int(combined["Quantity"].sum()) == 333

    print("\n== dimensions tab without Box Number ==")
    at.sidebar.toggle[1].set_value(False).run()
    assert not at.exception, at.exception
    dims = at.dataframe[1].value
    print("columns:", list(dims.columns))
    assert list(dims.columns) == ["Length", "Width", "Height"]

    print("\n== details sheet toggle ==")
    at.sidebar.toggle[2].set_value(True).run()
    assert not at.exception, at.exception
    print("still no errors:", not at.error)

    # The download payload is not exposed by AppTest, so check the writer itself.
    import openpyxl

    from packing_list import (
        build_dimensions,
        build_output,
        parse_packing_list,
        to_excel_bytes,
    )

    parsed = parse_packing_list(XLSX.read_bytes())
    contents = build_output(parsed, combine_duplicates=False)
    dimensions = build_dimensions(parsed)
    cases = [
        ((dimensions, None), ["Box Contents", "Box Dimensions"]),
        ((dimensions, parsed.lines), ["Box Contents", "Box Dimensions", "Details"]),
        ((None, None), ["Box Contents"]),
    ]
    for (dims_arg, detail_arg), expected in cases:
        wb = openpyxl.load_workbook(
            io.BytesIO(to_excel_bytes(contents, dims_arg, detail_arg))
        )
        print("sheets:", wb.sheetnames)
        assert wb.sheetnames == expected, (wb.sheetnames, expected)

    wb = openpyxl.load_workbook(
        io.BytesIO(to_excel_bytes(contents, dimensions, parsed.lines))
    )
    ws = wb["Box Contents"]
    assert ws["A2"].value == "250780202" and ws["A2"].number_format == "@"
    upcs = [row[0].value for row in ws.iter_rows(min_row=2, max_col=1)]
    assert all(isinstance(u, str) for u in upcs), "UPCs must be written as text"
    padded = sorted({u for u in upcs if u.startswith("0")})
    print("leading zeros kept as text in the download:", padded)
    assert padded == ["006020202", "050020202"], padded

    ws = wb["Box Dimensions"]
    header = [cell.value for cell in ws[1]]
    print("dimensions sheet header:", header)
    assert header == ["Box Number", "Length", "Width", "Height"]
    body = [row for row in ws.iter_rows(min_row=2, values_only=True)]
    print("first dimension rows:", body[:3], "row count:", len(body))
    assert len(body) == 28
    assert all(row[1:] == (31, 19, 14) for row in body)
    assert [row[0] for row in body] == list(range(1, 29))

    print("\napp renders cleanly")
finally:
    harness.unlink(missing_ok=True)
