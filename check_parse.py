"""Ad-hoc verification of the parser against a real Package Content List file."""

import sys

import pandas as pd

from packing_list import (
    build_dimensions,
    build_output,
    parse_packing_list,
    to_excel_bytes,
    upc_display_widths,
)

path = (
    sys.argv[1]
    if len(sys.argv) > 1
    else r"C:\Users\Administrator\Downloads\PMSH01 PO FBA19HZZ856W.xlsx"
)

parsed = parse_packing_list(path)
print("orders:", parsed.order_numbers)
print("cartons parsed:", len(parsed.cartons), "reported:", parsed.reported_carton_count)
print("qty parsed:", parsed.total_quantity, "reported:", parsed.reported_total_quantity)
print("weight parsed:", parsed.total_weight, "reported:", parsed.reported_net_weight)
print("item lines:", len(parsed.lines))
print("warnings:", parsed.warnings)

print("\nbox numbers:", [c.box_number for c in parsed.cartons])

raw = build_output(parsed, combine_duplicates=False)
combined = build_output(parsed, combine_duplicates=True)
print("\nrows raw:", len(raw), "rows combined:", len(combined))
print("qty raw:", raw["Quantity"].sum(), "qty combined:", combined["Quantity"].sum())

print("\n-- box 1 raw --")
print(raw[raw["Box Number"] == 1].to_string(index=False))

print("\n-- box 2 raw --")
print(raw[raw["Box Number"] == 2].to_string(index=False))

print("\n-- last box --")
print(raw[raw["Box Number"] == raw["Box Number"].max()].to_string(index=False))

print("\ndetail columns:", list(parsed.lines.columns))
print(parsed.lines.head(3).to_string(index=False))

dimensions = build_dimensions(parsed)
print("\n-- box dimensions --")
print("columns:", list(dimensions.columns), "dtypes:", dimensions.dtypes.to_dict())
print(dimensions.head(3).to_string(index=False))
print("rows:", len(dimensions))
print("unique L/W/H:", set(map(tuple, dimensions[["Length", "Width", "Height"]].values)))
assert len(dimensions) == parsed.reported_carton_count
assert list(dimensions.columns) == ["Box Number", "Length", "Width", "Height", "Weight"]
assert list(dimensions.loc[0, ["Length", "Width", "Height"]]) == [31, 19, 14]
assert dimensions["Weight"].notna().all(), "every box should have a Weight"
assert abs(float(dimensions.loc[0, "Weight"]) - 29.05) < 0.001
assert abs(float(dimensions["Weight"].sum()) - float(parsed.total_weight)) < 0.001

no_box = build_dimensions(parsed, include_box_number=False)
assert list(no_box.columns) == ["Length", "Width", "Height", "Weight"]

print("\n-- dtypes of the box contents table --")
print(raw.dtypes.to_dict())
assert all(pd.api.types.is_integer_dtype(raw[c]) for c in raw.columns), raw.dtypes
print("size dtype in detail:", parsed.lines["Size"].dtype)
assert pd.api.types.is_integer_dtype(parsed.lines["Size"])

# Box 1 must use real UPC barcodes, not Item numbers.
box1 = raw[raw["Box Number"] == 1]
assert box1["UPCs"].tolist()[0] == 673088044978
assert list(box1["Quantity"]) == [1, 3, 4, 2, 1, 1]
assert parsed.total_quantity == parsed.reported_total_quantity == 333
assert len(parsed.cartons) == 28

widths = upc_display_widths(parsed.lines["UPCs"])
print("zero-padded UPCs:", widths)

blob = to_excel_bytes(raw, dimensions, parsed.lines, upc_widths=widths)
print("\nexcel bytes:", len(blob))
with open("_verify_output.xlsx", "wb") as fh:
    fh.write(blob)

import openpyxl

wb = openpyxl.load_workbook("_verify_output.xlsx")
print("sheets:", wb.sheetnames)
assert wb.sheetnames == ["Box Contents", "Box Dimensions", "Details"]
for name in ("Box Contents", "Box Dimensions"):
    ws = wb[name]
    print(f"\n[{name}] dims:", ws.dimensions)
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        print(row)

ws = wb["Box Dimensions"]
assert ws["A1"].value == "Box Number"
assert [ws["B1"].value, ws["C1"].value, ws["D1"].value, ws["E1"].value] == [
    "Length",
    "Width",
    "Height",
    "Weight",
]
assert [ws["B2"].value, ws["C2"].value, ws["D2"].value] == [31, 19, 14]
assert abs(float(ws["E2"].value) - 29.05) < 0.001
assert all(isinstance(ws[f"{c}2"].value, (int, float)) for c in "BCDE")
assert ws.max_row == 29, ws.max_row

ws = wb["Box Contents"]
print("\n-- cell types on the contents sheet --")
for row in ws.iter_rows(min_row=2, max_col=3):
    upc, box, qty = row
    assert isinstance(upc.value, int), (upc.coordinate, upc.value, type(upc.value))
    assert isinstance(box.value, int) and isinstance(qty.value, int)
print("every UPC/Box/Quantity cell is a number")
print("first UPC cell:", ws["A2"].value)

detail = wb["Details"]
# UPCs, Box Number, Quantity, Item, Size, ..., Carton No
size_col = [cell.value for cell in detail[1]].index("Size") + 1
carton_col = [cell.value for cell in detail[1]].index("Carton No") + 1
print("detail size cell:", detail.cell(2, size_col).value)
print("detail carton cell:", repr(detail.cell(2, carton_col).value))
assert isinstance(detail.cell(2, size_col).value, (int, float))
assert isinstance(detail.cell(2, carton_col).value, str)
