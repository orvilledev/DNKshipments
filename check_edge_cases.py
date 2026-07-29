"""Edge-case checks: page-split cartons, multiple sheets, bad input."""

import io

import openpyxl
import pandas as pd

from packing_list import (
    PackingListError,
    build_dimensions,
    build_output,
    parse_dimensions,
    parse_packing_list,
    to_excel_bytes,
)


def write(rows, sheets=None):
    """rows: list of (row, col, value) using 1-based indexes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "#1"
    for r, c, v in rows:
        ws.cell(row=r, column=c, value=v)
    for name, extra in (sheets or {}).items():
        ws2 = wb.create_sheet(name)
        for r, c, v in extra:
            ws2.cell(row=r, column=c, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def carton_block(start, carton_no, seq, total, items):
    rows = [
        (start, 1, "Carton No:"),
        (start, 5, carton_no),
        (start, 13, f"Carton: {seq} of {total}"),
        (start, 16, "Dimensions: 31x19x14"),
        (start + 1, 25, "Unit of Measure Code"),
        (start + 2, 3, "Item"),
        (start + 2, 7, "Size"),
        (start + 2, 14, "Description"),
        (start + 2, 21, "Quantity"),
    ]
    r = start + 3
    for upc, size, desc, qty in items:
        rows += [(r, 3, upc), (r, 8, size), (r, 14, desc), (r, 21, qty), (r, 25, "PAIR")]
        r += 1
    rows += [(r, 17, "Carton Total:"), (r, 21, sum(i[3] for i in items))]
    return rows, r + 2


print("== case 1: carton continued on a second print page ==")
rows = [(1, 1, "Package Content List"), (2, 1, "Order No."), (2, 5, "SO-1")]
block, nxt = carton_block(4, "L001", 1, 2, [(111111111, "40", "A", 2)])
# same carton header repeated after a page break, with more items
rows += block[:-2]
rows += [
    (nxt - 2, 19, "Page 2 / 2"),
    (nxt - 1, 1, "Carton No:"),
    (nxt - 1, 5, "L001"),
    (nxt - 1, 13, "Carton: 1 of 2"),
    (nxt, 3, "Item"),
    (nxt, 7, "Size"),
    (nxt, 14, "Description"),
    (nxt, 21, "Quantity"),
    (nxt + 1, 3, 222222222),
    (nxt + 1, 8, "41"),
    (nxt + 1, 14, "B"),
    (nxt + 1, 21, 3),
    (nxt + 2, 17, "Carton Total:"),
    (nxt + 2, 21, 5),
]
block2, nxt2 = carton_block(nxt + 4, "L002", 2, 2, [(333333333, "42", "C", 4)])
rows += block2
rows += [(nxt2, 2, "Total Net Weight:"), (nxt2, 10, 10.5), (nxt2, 14, "Total # of Cartons:"), (nxt2, 18, 2), (nxt2, 23, "Total Quantity:"), (nxt2, 26, 9)]

parsed = parse_packing_list(write(rows))
print("cartons:", len(parsed.cartons), "boxes:", [c.box_number for c in parsed.cartons])
print("qty:", parsed.total_quantity, "reported:", parsed.reported_total_quantity)
print("warnings:", parsed.warnings)
print(build_output(parsed).to_string(index=False))
assert len(parsed.cartons) == 2, "page-split carton should stay one box"
assert parsed.total_quantity == 9
assert not parsed.warnings

print("\n== case 2: cartons spread across two worksheets ==")
b1, e1 = carton_block(1, "L001", 1, 2, [(111111111, "40", "A", 2)])
b2, e2 = carton_block(1, "L002", 2, 2, [(222222222, "41", "B", 3)])
b2 += [(e2, 2, "Total Net Weight:"), (e2, 14, "Total # of Cartons:"), (e2, 18, 2), (e2, 23, "Total Quantity:"), (e2, 26, 5)]
parsed = parse_packing_list(write(b1, sheets={"#2": b2}))
print("boxes:", [(c.box_number, c.carton_no, c.sheet) for c in parsed.cartons])
print("qty:", parsed.total_quantity, "warnings:", parsed.warnings)
assert [c.box_number for c in parsed.cartons] == [1, 2]
assert parsed.total_quantity == 5

print("\n== case 3: carton total disagrees with its lines ==")
b, e = carton_block(1, "L001", 1, 1, [(111111111, "40", "A", 2)])
b = [(r, c, 99 if (r, c) == (e - 2, 21) else v) for r, c, v in b]
parsed = parse_packing_list(write(b))
print("warnings:", parsed.warnings)
assert parsed.warnings and "Carton Total" in parsed.warnings[0]

print("\n== case 4: no cartons at all ==")
try:
    parse_packing_list(write([(1, 1, "Some other report"), (3, 1, "Hello")]))
except PackingListError as exc:
    print("raised:", exc)
else:
    raise AssertionError("expected PackingListError")

print("\n== case 5: not an xlsx ==")
try:
    parse_packing_list(b"this is not a workbook")
except PackingListError as exc:
    print("raised ok")
else:
    raise AssertionError("expected PackingListError")

print("\n== case 6: missing 'Carton: N of M' label falls back to sequence ==")
rows = []
r = 1
for i, carton in enumerate(["L001", "L002"], start=1):
    rows += [
        (r, 1, "Carton No:"),
        (r, 5, carton),
        (r + 1, 3, "Item"),
        (r + 1, 7, "Size"),
        (r + 1, 14, "Description"),
        (r + 1, 21, "Quantity"),
        (r + 2, 3, 100000000 + i),
        (r + 2, 8, "40"),
        (r + 2, 14, "X"),
        (r + 2, 21, i),
    ]
    r += 5
parsed = parse_packing_list(write(rows))
print("boxes:", [(c.box_number, c.carton_no) for c in parsed.cartons], "warnings:", parsed.warnings)
assert [c.box_number for c in parsed.cartons] == [1, 2]

print("\n== case 7: dimension strings ==")
expected = {
    "31x19x14": (31, 19, 14),
    "31 x 19 x 14": (31, 19, 14),
    "31X19X14": (31, 19, 14),
    "31\u00d719\u00d714": (31, 19, 14),
    "31*19*14": (31, 19, 14),
    "31x19x14 in": (31, 19, 14),
    "31.5x19x14.25": (31.5, 19, 14.25),
    "31,5x19x14": (31.5, 19, 14),
    "L31 W19 H14": (31, 19, 14),
    "31x19": (None, None, None),
    "": (None, None, None),
}
for text, want in expected.items():
    got = parse_dimensions(text)
    print(f"{text!r:16} -> {got}")
    assert got == want, (text, got, want)

print("\n== case 8: a box with no dimensions is flagged, others still parse ==")
b1, e1 = carton_block(1, "L001", 1, 2, [(111111111, "40", "A", 2)])
rows = [(r, c, v) for r, c, v in b1 if not (c == 16 and str(v).startswith("Dimensions"))]
b2, e2 = carton_block(e1, "L002", 2, 2, [(222222222, "41", "B", 3)])
parsed = parse_packing_list(write(rows + b2))
print("warnings:", parsed.warnings)
dims = build_dimensions(parsed)
print(dims.to_string(index=False))
assert len(parsed.warnings) == 1 and "dimensions" in parsed.warnings[0]
assert dims["Length"].isna().tolist() == [True, False]
assert dims.loc[1, "Length"] == 31 and dims.loc[1, "Height"] == 14

print("\n== case 9: decimal dimensions stay decimal in the table ==")
b, _ = carton_block(1, "L001", 1, 1, [(111111111, "40", "A", 1)])
b = [(r, c, "Dimensions: 31.5x19x14" if c == 16 else v) for r, c, v in b]
dims = build_dimensions(parse_packing_list(write(b)))
print(dims.to_string(index=False))
assert dims.loc[0, "Length"] == 31.5 and dims.loc[0, "Width"] == 19

print("\n== case 10: numbers are numbers, non-numeric codes stay text ==")
b, _ = carton_block(1, "L001", 1, 1, [(111111111, "40", "A", 2), (222222222, "41", "B", 1)])
parsed = parse_packing_list(write(b))
out = build_output(parsed)
print("dtypes:", out.dtypes.to_dict())
assert all(pd.api.types.is_integer_dtype(out[c]) for c in out.columns), out.dtypes

rows = []
r = 1
rows += [
    (r, 1, "Carton No:"),
    (r, 5, "L001"),
    (r, 13, "Carton: 1 of 1"),
    (r, 16, "Dimensions: 31x19x14"),
    (r + 1, 3, "Item"),
    (r + 1, 7, "Size"),
    (r + 1, 14, "Description"),
    (r + 1, 21, "Quantity"),
    (r + 2, 3, "ABC-123"),
    (r + 2, 8, "40"),
    (r + 2, 14, "A"),
    (r + 2, 21, 2),
]
alpha = build_output(parse_packing_list(write(rows)))
print("alphanumeric UPCs dtype:", alpha["UPCs"].dtype, "->", alpha["UPCs"].tolist())
assert alpha["UPCs"].dtype == object and alpha.loc[0, "UPCs"] == "ABC-123"

wb = openpyxl.load_workbook(io.BytesIO(to_excel_bytes(alpha)))
cell = wb["Box Contents"]["A2"]
print("excel cell:", repr(cell.value), cell.number_format)
assert cell.value == "ABC-123" and cell.number_format == "@"

print("\n== case 11: UPC column preferred over Item, and sticky across cartons ==")
# Carton 1 header includes UPC (col P=16). Carton 2 omits the UPC label but
# still puts UPC values in column P — matching the real Dansko export.
rows = [
    (1, 1, "Order No."),
    (1, 5, "SO-UPC"),
    (2, 1, "Carton No:"),
    (2, 5, "L001"),
    (2, 13, "Carton: 1 of 2"),
    (2, 16, "Dimensions: 31x19x14"),
    (3, 25, "Unit of Measure Code"),
    (4, 3, "Item"),
    (4, 7, "Size"),
    (4, 16, "UPC"),
    (4, 17, "Description"),
    (4, 21, "Quantity"),
    (4, 28, "Weight"),
    (4, 29, "TOTAL WEIGHT"),
    (5, 3, 906020202),
    (5, 8, 38),
    (5, 16, "673088939052"),
    (5, 17, "Professional Black Tooled"),
    (5, 21, 4),
    (5, 25, "PAIR"),
    (6, 17, "Carton Total:"),
    (6, 21, 4),
    (8, 1, "Carton No:"),
    (8, 5, "L002"),
    (8, 13, "Carton: 2 of 2"),
    (8, 16, "Dimensions: 27x19x14"),
    (9, 3, "Item"),
    (9, 7, "Size"),
    (9, 17, "Description"),
    (9, 21, "Quantity"),
    (10, 3, 238581212),
    (10, 8, 39),
    (10, 16, "673088460655"),
    (10, 17, "Ingrid Honey Distressed"),
    (10, 21, 3),
    (11, 17, "Carton Total:"),
    (11, 21, 3),
    (13, 2, "Total Net Weight:"),
    (13, 14, "Total # of Cartons:"),
    (13, 18, 2),
    (13, 23, "Total Quantity:"),
    (13, 26, 7),
]
parsed = parse_packing_list(write(rows))
out = build_output(parsed, combine_duplicates=False)
print(out.to_string(index=False))
print("detail Item column:", parsed.lines["Item"].tolist())
assert parsed.total_quantity == 7 and len(parsed.cartons) == 2
assert out["UPCs"].tolist() == [673088939052, 673088460655]
assert parsed.lines["Item"].tolist() == ["906020202", "238581212"]
assert list(build_dimensions(parsed)["Length"]) == [31, 27]
assert not parsed.warnings

print("\nall edge cases passed")
